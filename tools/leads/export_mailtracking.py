#!/usr/bin/env python3
"""Traegt Kontakte aus leads.csv in das Mail-Tracking-Blatt ein.

Schreibt ausschliesslich Nr, Firma, Ansprechpartner, E-Mail und Notiz in das
Blatt 'Mails'. Versanddaten, Variante und Antwortart bleiben leer — die traegt
der Mensch ein. Formeln, Dropdowns und Formatierung werden nicht angefasst.

    python export_mailtracking.py --csv leads.csv --xlsx MailTracking-ELEVO.xlsx
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

import openpyxl

# Zeile 2 ist die Beispielzeile der Vorlage; die Auswertung rechnet ab Zeile 3.
ERSTE_ZEILE = 3
LETZTE_ZEILE = 502

# Werte, die eine Firmierung statt einer Person bezeichnen
FIRMA_MUSTER = re.compile(
    r"\b(GmbH|mbH|AG|KG|UG|SE|e\.K\.|Verwaltungs|Beteiligungs|Holding|Gesellschaft)\b",
    re.IGNORECASE,
)


def primaeradresse(zeile: dict[str, str]) -> str:
    """Funktionspostfach bevorzugen — die uebliche Adresse fuer die Erstansprache."""
    paare = list(zip(zeile["email"].split("|"), zeile["email_typ"].split("|")))
    for adresse, typ in paare:
        if typ == "funktion":
            return adresse
    return paare[0][0]


def kontakte_aus_csv(pfad: Path) -> list[dict]:
    with pfad.open(encoding="utf-8-sig") as datei:
        zeilen = [z for z in csv.DictReader(datei, delimiter=";") if z["email"]]

    # Nach Branche, Ort, Firma sortieren, damit sich Versandwellen bilden lassen
    zeilen.sort(key=lambda z: (z["branche_suchbegriff"], z["ort"], z["firma"]))

    kontakte: list[dict] = []
    gesehen: dict[str, dict] = {}
    for zeile in zeilen:
        adresse = primaeradresse(zeile)
        if adresse in gesehen:
            # Sammeladresse mehrerer Standorte — nur einmal anschreiben
            gesehen[adresse]["weitere"] += 1
            continue
        eintrag = {"quelle": zeile, "email": adresse, "weitere": 0}
        gesehen[adresse] = eintrag
        kontakte.append(eintrag)
    return kontakte


def eintragen(kontakte: list[dict], xlsx: Path) -> int:
    kapazitaet = LETZTE_ZEILE - ERSTE_ZEILE + 1
    if len(kontakte) > kapazitaet:
        raise SystemExit(
            f"{len(kontakte)} Kontakte passen nicht in {kapazitaet} Zeilen. "
            "Entweder die Vorlage verlaengern oder die Liste vorher filtern."
        )

    wb = openpyxl.load_workbook(xlsx)
    ws = wb["Mails"]

    for versatz, eintrag in enumerate(kontakte):
        quelle = eintrag["quelle"]
        zeile = ERSTE_ZEILE + versatz

        person = quelle["ansprechpartner"]
        if person and FIRMA_MUSTER.search(person):
            person = ""  # Firmierung ist kein Ansprechpartner

        notiz = f"{quelle['branche_suchbegriff']} · {quelle['ort']}"
        if eintrag["weitere"]:
            notiz += f" · Sammeladresse, {eintrag['weitere']} weitere Standorte"

        ws.cell(row=zeile, column=1, value=versatz + 1)
        ws.cell(row=zeile, column=2, value=quelle["firma"])
        ws.cell(row=zeile, column=3, value=person)
        ws.cell(row=zeile, column=4, value=eintrag["email"])
        ws.cell(row=zeile, column=13, value=notiz)

    wb.save(xlsx)
    return len(kontakte)


def main() -> int:
    basis = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--csv", default=str(basis / "leads.csv"))
    parser.add_argument("--xlsx", required=True, help="Mail-Tracking-Mappe (wird ueberschrieben)")
    args = parser.parse_args()

    kontakte = kontakte_aus_csv(Path(args.csv))
    anzahl = eintragen(kontakte, Path(args.xlsx))
    sammel = sum(1 for k in kontakte if k["weitere"])
    print(
        f"{anzahl} Kontakte in Zeilen {ERSTE_ZEILE}-{ERSTE_ZEILE + anzahl - 1} "
        f"eingetragen ({sammel} Sammeladressen zusammengefasst)",
        file=sys.stderr,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
